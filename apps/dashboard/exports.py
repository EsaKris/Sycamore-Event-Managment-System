"""
Turns (columns, rows) — as produced by apps/dashboard/reports.py — into
a downloadable file. Three formats per spec: PDF, Excel, CSV.
"""

import csv
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def export_csv(title: str, columns: list, rows: list) -> HttpResponse:
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{title}.csv"'
    writer = csv.DictWriter(response, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return response


def export_excel(title: str, columns: list, rows: list) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_fill = PatternFill(start_color='1C1B18', end_color='1C1B18', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ''))

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(r.get(col_name, ''))) for r in rows]) if rows else len(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return response


def export_pdf(title: str, columns: list, rows: list) -> HttpResponse:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 8)]

    if not rows:
        elements.append(Paragraph('No data for this report.', styles['Normal']))
    else:
        # Truncate to a printable page count — for anything larger, Excel/CSV is the right tool.
        printable_rows = rows[:500]
        data = [columns] + [[str(r.get(c, ''))[:60] for c in columns] for r in printable_rows]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1C1B18')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D8D4C8')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F6F3')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)
        if len(rows) > 500:
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(
                f'Showing the first 500 of {len(rows)} rows. Export as Excel or CSV for the full dataset.',
                styles['Italic'],
            ))

    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title}.pdf"'
    return response


EXPORTERS = {'csv': export_csv, 'xlsx': export_excel, 'pdf': export_pdf}
